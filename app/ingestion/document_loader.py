"""
Document loader — supports PDF, DOCX, TXT, MD, CSV, XLSX, HTML, PPTX.
Each loader returns a list[Document] with rich metadata attached.
"""
from __future__ import annotations

import hashlib
from pathlib import Path
from typing import List

from langchain_core.documents import Document

from app import config
from app.utils.logger import get_logger

logger = get_logger(__name__)


def _file_hash(path: Path) -> str:
    """SHA-256 of file content — used for deduplication."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _base_metadata(path: Path) -> dict:
    return {
        "source": str(path.resolve()),
        "filename": path.name,
        "extension": path.suffix.lower(),
        "file_hash": _file_hash(path),
    }


# ──────────────────────────────────────────────
# Per-format loaders
# ──────────────────────────────────────────────

def _load_pdf(path: Path) -> List[Document]:
    from langchain_community.document_loaders import PyMuPDFLoader

    loader = PyMuPDFLoader(str(path))
    docs = loader.load()
    for doc in docs:
        doc.metadata.update(_base_metadata(path))
    logger.debug("Loaded %d pages from PDF: %s", len(docs), path.name)
    return docs


def _load_docx(path: Path) -> List[Document]:
    from langchain_community.document_loaders import Docx2txtLoader

    loader = Docx2txtLoader(str(path))
    docs = loader.load()
    for doc in docs:
        doc.metadata.update(_base_metadata(path))
    logger.debug("Loaded DOCX: %s", path.name)
    return docs


def _load_txt_or_md(path: Path) -> List[Document]:
    from langchain_community.document_loaders import TextLoader

    loader = TextLoader(str(path), encoding="utf-8", autodetect_encoding=True)
    docs = loader.load()
    for doc in docs:
        doc.metadata.update(_base_metadata(path))
    return docs


def _load_csv(path: Path) -> List[Document]:
    from langchain_community.document_loaders.csv_loader import CSVLoader

    loader = CSVLoader(file_path=str(path), encoding="utf-8")
    docs = loader.load()
    for doc in docs:
        doc.metadata.update(_base_metadata(path))
    logger.debug("Loaded %d rows from CSV: %s", len(docs), path.name)
    return docs


def _load_excel(path: Path) -> List[Document]:
    """Converts each sheet's rows into individual Documents."""
    import openpyxl
    from langchain_core.documents import Document

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    docs: List[Document] = []
    meta = _base_metadata(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_text = " | ".join(
                f"{h}: {v}" for h, v in zip(headers, row) if v is not None
            )
            if row_text.strip():
                docs.append(
                    Document(
                        page_content=row_text,
                        metadata={**meta, "sheet": sheet_name},
                    )
                )
    wb.close()
    logger.debug("Loaded %d rows from Excel: %s", len(docs), path.name)
    return docs


def _load_html(path: Path) -> List[Document]:
    from langchain_community.document_loaders import UnstructuredHTMLLoader

    loader = UnstructuredHTMLLoader(str(path))
    docs = loader.load()
    for doc in docs:
        doc.metadata.update(_base_metadata(path))
    return docs


def _load_pptx(path: Path) -> List[Document]:
    from langchain_community.document_loaders import UnstructuredPowerPointLoader

    loader = UnstructuredPowerPointLoader(str(path))
    docs = loader.load()
    for doc in docs:
        doc.metadata.update(_base_metadata(path))
    return docs


# ──────────────────────────────────────────────
# Dispatcher
# ──────────────────────────────────────────────

_LOADER_MAP = {
    ".pdf": _load_pdf,
    ".docx": _load_docx,
    ".doc": _load_docx,
    ".txt": _load_txt_or_md,
    ".md": _load_txt_or_md,
    ".csv": _load_csv,
    ".xlsx": _load_excel,
    ".xls": _load_excel,
    ".html": _load_html,
    ".htm": _load_html,
    ".pptx": _load_pptx,
}


def load_document(path: Path) -> List[Document]:
    """Load a single file; returns an empty list if the format is unsupported."""
    ext = path.suffix.lower()
    loader_fn = _LOADER_MAP.get(ext)
    if loader_fn is None:
        logger.warning("Unsupported file type skipped: %s", path)
        return []
    try:
        return loader_fn(path)
    except Exception as exc:
        logger.error("Failed to load %s — %s", path, exc, exc_info=True)
        return []


def load_directory(directory: Path) -> List[Document]:
    """
    Recursively walk *directory* and load every supported file.
    Returns the combined list of Documents.
    """
    directory = Path(directory)
    if not directory.exists():
        raise FileNotFoundError(f"Documents directory not found: {directory}")

    all_docs: List[Document] = []
    files = [
        p
        for p in directory.rglob("*")
        if p.is_file() and p.suffix.lower() in config.SUPPORTED_EXTENSIONS
    ]

    logger.info("Found %d supported files in '%s'", len(files), directory)
    for file_path in files:
        docs = load_document(file_path)
        all_docs.extend(docs)

    logger.info("Total raw document chunks loaded: %d", len(all_docs))
    return all_docs
